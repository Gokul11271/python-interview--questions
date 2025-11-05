import openpyxl
import random
import os
from openpyxl import load_workbook

# === Template and output settings ===
template_path = r"C:\Users\Admin\Downloads\EXCEL BATCH REPORT  07.06.2025.xlsx"
output_folder = r"C:\Users\Admin\Downloads\Generated_Bills"

os.makedirs(output_folder, exist_ok=True)

# === Value ranges ===
RANGE_40MM = (1439, 1445)
RANGE_M_SA = (523, 530)
RANGE_CEM2 = (3671, 3674)
RANGE_WATER = (133, 140)

def modify_excel(template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb.active

    # Detect header row & column indexes
    header_row_index = None
    col_40mm = col_msa = col_cem2 = col_water = None

    for row in ws.iter_rows(min_row=1, max_row=40):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = str(cell.value).strip().upper()
                if val == "40MM":
                    col_40mm = cell.column
                    header_row_index = cell.row
                elif val == "M SA":
                    col_msa = cell.column
                elif val == "CEM2":
                    col_cem2 = cell.column
                elif val == "WATER":
                    col_water = cell.column

    if not all([col_40mm, col_msa, col_cem2, col_water]):
        print("⚠️ Could not detect all required columns (40MM, M SA, CEM2, WATER).")
        return

    # Prepare sums for totals
    total_40mm = 0
    total_msa = 0
    total_cem2 = 0
    total_water = 0

    # Modify numeric rows
    for row in ws.iter_rows(min_row=header_row_index + 1):
        c40 = ws.cell(row=row[0].row, column=col_40mm)
        cma = ws.cell(row=row[0].row, column=col_msa)
        cce = ws.cell(row=row[0].row, column=col_cem2)
        cwa = ws.cell(row=row[0].row, column=col_water)

        if all(isinstance(c.value, (int, float)) for c in [c40, cma, cce, cwa] if c.value not in [None, ""]):
            new_40 = random.randint(*RANGE_40MM)
            new_msa = random.randint(*RANGE_M_SA)
            new_cem2 = random.randint(*RANGE_CEM2)
            new_water = random.randint(*RANGE_WATER)

            c40.value = new_40
            cma.value = new_msa
            cce.value = new_cem2
            cwa.value = new_water

            total_40mm += new_40
            total_msa += new_msa
            total_cem2 += new_cem2
            total_water += new_water

    # Update totals below (Total Set Weight / Total Actual)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = str(cell.value).strip().upper()
                if "TOTAL SET WEIGHT" in val:
                    # row below contains numeric totals
                    target_row = cell.row + 1
                    ws.cell(row=target_row, column=col_40mm, value=0)
                    ws.cell(row=target_row, column=col_msa, value=total_msa)
                    ws.cell(row=target_row, column=col_cem2, value=total_cem2)
                    ws.cell(row=target_row, column=col_water, value=total_water)
                elif "TOTAL ACTUAL" in val:
                    target_row = cell.row + 1
                    ws.cell(row=target_row, column=col_40mm, value=0)
                    ws.cell(row=target_row, column=col_msa, value=total_msa + random.randint(-5, 5))
                    ws.cell(row=target_row, column=col_cem2, value=total_cem2 + random.randint(-5, 5))
                    ws.cell(row=target_row, column=col_water, value=total_water + random.randint(-3, 3))

    wb.save(output_path)
    wb.close()

# === Generate 10 Excel bills ===
for i in range(1, 11):
    output_path = os.path.join(output_folder, f"Billing_{i:02}.xlsx")
    modify_excel(template_path, output_path)
    print(f"✅ Generated: {output_path}")

print("\n🎯 All 10 formatted billing sheets created successfully with updated totals!")
